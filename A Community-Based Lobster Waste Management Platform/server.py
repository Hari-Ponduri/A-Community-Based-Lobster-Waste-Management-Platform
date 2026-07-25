"""
ShellCycle Backend Server
Flask API that uses an Excel file (database.xlsx) as a simple database.
Endpoints:
  POST /api/restaurants  — Append a new restaurant row
  GET  /api/restaurants  — Return all saved restaurants as JSON
"""

from flask import Flask, request, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

app = Flask(__name__)
CORS(app)  # Allow cross-origin requests from the frontend

DB_FILE = "database.xlsx"
SHEET_NAME = "Restaurants"
HEADERS = ["ID", "Name", "Shell Type", "Weekly Kg", "Storage", "Pickup Window", "Location", "Registered At"]


def init_db():
    """Create database.xlsx with headers if it doesn't exist."""
    if not os.path.exists(DB_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(HEADERS)
        wb.save(DB_FILE)
        print(f"[ShellCycle] Created {DB_FILE} with headers.")
    else:
        print(f"[ShellCycle] Using existing {DB_FILE}.")


def load_workbook_safe():
    """Load the workbook, re-initialising if it's corrupted/missing."""
    try:
        wb = openpyxl.load_workbook(DB_FILE)
        if SHEET_NAME not in wb.sheetnames:
            ws = wb.create_sheet(SHEET_NAME)
            ws.append(HEADERS)
            wb.save(DB_FILE)
        return wb
    except Exception:
        init_db()
        return openpyxl.load_workbook(DB_FILE)


# ──────────────────────────────────────────────
# POST /api/restaurants  — Save a new restaurant
# ──────────────────────────────────────────────
@app.route("/api/restaurants", methods=["POST"])
def add_restaurant():
    data = request.get_json(force=True)

    required = ["id", "name", "shellType", "weeklyKg", "storage", "pickupWindow", "location"]
    for field in required:
        if field not in data:
            return jsonify({"error": f"Missing field: {field}"}), 400

    row = [
        str(data["id"]),
        str(data["name"]),
        str(data["shellType"]),
        float(data["weeklyKg"]),
        str(data["storage"]),
        str(data["pickupWindow"]),
        str(data["location"]),
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]

    wb = load_workbook_safe()
    ws = wb[SHEET_NAME]
    ws.append(row)
    wb.save(DB_FILE)

    print(f"[ShellCycle] Saved restaurant: {data['name']}")
    return jsonify({"message": "Restaurant saved successfully", "restaurant": data}), 201


# ──────────────────────────────────────────────
# GET /api/restaurants  — Fetch all restaurants
# ──────────────────────────────────────────────
@app.route("/api/restaurants", methods=["GET"])
def get_restaurants():
    wb = load_workbook_safe()
    ws = wb[SHEET_NAME]

    restaurants = []
    rows = list(ws.iter_rows(values_only=True))

    # Skip the header row (index 0)
    for row in rows[1:]:
        if row[0] is None:
            continue
        restaurants.append({
            "id":           str(row[0]),
            "name":         str(row[1]),
            "shellType":    str(row[2]),
            "weeklyKg":     float(row[3]),
            "storage":      str(row[4]),
            "pickupWindow": str(row[5]),
            "location":     str(row[6]),
            "registeredAt": str(row[7]) if row[7] else "",
        })

    print(f"[ShellCycle] Returning {len(restaurants)} restaurants.")
    return jsonify({"restaurants": restaurants, "count": len(restaurants)}), 200


# ──────────────────────────────────────────────
# Health check
# ──────────────────────────────────────────────
@app.route("/api/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "db": DB_FILE}), 200


if __name__ == "__main__":
    init_db()
    print("=" * 50)
    print("  ShellCycle Backend Server")
    print("  Running at http://localhost:5000")
    print("  POST /api/restaurants — Save restaurant")
    print("  GET  /api/restaurants — List all restaurants")
    print("=" * 50)
    app.run(host="0.0.0.0", port=5000, debug=True, use_reloader=False)
